#!/usr/bin/env python
"""
Generate a data input template for new crawl projects.

Creates a sample Excel file with the correct structure and example data.
Users can copy this file and fill in their own data.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import config
sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

PROJECT_ROOT = config.PROJECT_ROOT


def create_template() -> None:
    """Create a data input template file."""
    
    template_path = PROJECT_ROOT / "docs" / "data_template.xlsx"
    template_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create sample data with Vietnamese column names
    sample_data = {
        "Mã phòng ban": [1001, 1002, 1003, 1004, 1005],
        "Tên phòng ban": [
            "Chi nhánh TP. Hồ Chí Minh",
            "Chi nhánh Hà Nội",
            "Chi nhánh Đà Nẵng",
            "Chi nhánh Cần Thơ",
            "Chi nhánh Hải Phòng",
        ],
        "KINH ĐỘ": [106.6296, 105.8342, 108.2022, 105.7778, 106.6841],
        "VĨ ĐỘ": [10.7769, 21.0285, 16.0544, 10.0341, 20.8448],
        "Tỉnh": ["Hồ Chí Minh", "Hà Nội", "Đà Nẵng", "Cần Thơ", "Hải Phòng"],
    }
    
    df = pd.DataFrame(sample_data)
    
    # Write to Excel
    df.to_excel(template_path, sheet_name="Data", index=False)
    
    # Format the Excel file
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Format header row
    for cell in ws[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
    
    # Format data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.column in [3, 4]:  # KINH ĐỘ, Vĩ ĐỘ
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 25
    
    # Add instruction sheet
    ws_instruction = wb.create_sheet("Hướng dẫn", 0)
    
    instructions = [
        ["HƯỚNG DẪN ĐIỀN DỮ LIỆU", ""],
        ["", ""],
        ["Cột dữ liệu bắt buộc:", ""],
        ["1. Mã phòng ban", "Số hiệu duy nhất cho mỗi phòng/chi nhánh (VD: 1001, 1002)"],
        ["2. Tên phòng ban", "Tên đầy đủ của phòng/chi nhánh"],
        ["3. KINH ĐỘ", "Tọa độ kinh độ (VD: 106.6296)"],
        ["4. VĨ ĐỘ", "Tọa độ vĩ độ (VD: 10.7769)"],
        ["5. Tỉnh", "Tỉnh/Thành phố (VD: Hồ Chí Minh, Hà Nội)"],
        ["", ""],
        ["Lưu ý:", ""],
        ["• Mỗi hàng là một phòng/chi nhánh", ""],
        ["• Không để trống bất kỳ cột nào", ""],
        ["• Tọa độ phải là số thực (VD: 106.6296, không phải 106°38')"],
        ["• Đặt tên tỉnh theo chuẩn hiện hành", ""],
        ["", ""],
        ["Ví dụ dữ liệu:", ""],
        ["Mã phòng ban", "Tên phòng ban", "KINH ĐỘ", "VĨ ĐỘ", "Tỉnh"],
        ["1001", "Chi nhánh TP. Hồ Chí Minh", "106.6296", "10.7769", "Hồ Chí Minh"],
        ["1002", "Chi nhánh Hà Nội", "105.8342", "21.0285", "Hà Nội"],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instruction.cell(row=row_idx, column=col_idx, value=value)
            
            # Format title row
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Format section headers
            elif row_idx in [3, 9, 15]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Format example table
            elif row_idx == 16:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
    
    ws_instruction.column_dimensions["A"].width = 25
    ws_instruction.column_dimensions["B"].width = 45
    
    wb.save(template_path)
    
    print(f"✅ Template file created: {template_path}")
    print(f"\n📋 Sheet 1: 'Hướng dẫn' - Instructions for users")
    print(f"📋 Sheet 2: 'Data' - Sample data to modify and use")
    print(f"\n💡 Users should:")
    print(f"   1. Download this file")
    print(f"   2. Replace example data with their own")
    print(f"   3. Save as: data_YYYYMMDD/input/branches_YYYYMMDD.xlsx")


if __name__ == "__main__":
    try:
        create_template()
        print("\n✓ Done!")
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)
