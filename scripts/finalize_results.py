#!/usr/bin/env python
"""Merge crawl outputs, split distance/time, and write final diagnostics."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# Import config
sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

PROJECT_ROOT = config.PROJECT_ROOT

# Use config for column names and regex patterns
DISTANCE_COL = config.DISTANCE_COL
KM_COL = config.KM_COL
TIME_COL = config.TIME_COL
DISTANCE_RE = config.DISTANCE_REGEX
ERROR_RE = config.ERROR_KEYWORDS

# Regex for parsing output directories and files
import re
OUTPUT_DIR_RE = re.compile(r"output_part_(\d+)$")
OUTPUT_FILE_RE = re.compile(r"ket_qua_tu_(\d+)_den_(\d+)\.xlsx$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--version",
        type=str,
        default=config.DATA_VERSION,
        help=f"Data version (YYYYMMDD format). Default: {config.DATA_VERSION}",
    )
    
    # Build defaults from config
    paths = config.get_data_paths(config.DATA_VERSION)
    default_source = paths["crawl_pairs_file"]
    default_part_folder = config.PART_FOLDER
    default_final = paths["final_file"]
    default_remaining = paths["failed_file"]
    default_summary = paths["summary_file"]
    
    parser.add_argument("--source", type=Path, default=default_source)
    parser.add_argument("--part-folder", type=Path, default=default_part_folder)
    parser.add_argument("--final", type=Path, default=default_final)
    parser.add_argument("--remaining", type=Path, default=default_remaining)
    parser.add_argument("--summary", type=Path, default=default_summary)
    
    args = parser.parse_args()
    
    # If version changed, update paths
    if args.version != config.DATA_VERSION:
        paths = config.get_data_paths(args.version)
        if args.source == default_source:
            args.source = paths["crawl_pairs_file"]
        if args.final == default_final:
            args.final = paths["final_file"]
        if args.remaining == default_remaining:
            args.remaining = paths["failed_file"]
        if args.summary == default_summary:
            args.summary = paths["summary_file"]
    
    return args


def parse_distance_time(value: object) -> tuple[float | None, str | None]:
    if pd.isna(value):
        return None, None

    text = str(value).strip()
    if not text or ERROR_RE.search(text):
        return None, None

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for idx in range(len(lines) - 1, -1, -1):
        match = DISTANCE_RE.search(lines[idx])
        if not match:
            continue

        number = float(match.group("num").replace(",", "."))
        unit = match.group("unit").lower()
        km = number if unit == "km" else number / 1000
        time_text = " ".join(lines[:idx] + lines[idx + 1 :]).strip() or None
        return round(km, 2), time_text

    return None, " ".join(lines) or None


def read_output_files(part_folder: Path, source_rows: int) -> pd.DataFrame:
    records: list[pd.DataFrame] = []

    for path in sorted(part_folder.glob("output_part_*/ket_qua_tu_*_den_*.xlsx")):
        dir_match = OUTPUT_DIR_RE.search(path.parent.name)
        file_match = OUTPUT_FILE_RE.search(path.name)
        if not dir_match or not file_match:
            continue

        part_id = int(dir_match.group(1))
        start = int(file_match.group(1))
        end = int(file_match.group(2))
        df = pd.read_excel(path).reset_index(drop=True)

        if "global_index" not in df.columns:
            df["global_index"] = (part_id - 1) * 100 + start + df.index

        df = df[(df["global_index"] >= 0) & (df["global_index"] < source_rows)].copy()
        df["_source_output_file"] = str(path.relative_to(PROJECT_ROOT))
        df["_range_width"] = end - start
        records.append(df[["global_index", DISTANCE_COL, "_source_output_file", "_range_width"]])

    if not records:
        return pd.DataFrame(
            columns=["global_index", DISTANCE_COL, "_source_output_file", "_range_width"]
        )

    results = pd.concat(records, ignore_index=True)
    text = results[DISTANCE_COL].astype("string").str.strip()
    results["_has_value"] = results[DISTANCE_COL].notna() & text.notna() & text.ne("")
    results["_is_error"] = text.str.contains(ERROR_RE, na=False)
    return results.sort_values(
        ["global_index", "_has_value", "_is_error", "_range_width"],
        ascending=[True, False, True, False],
    )


def main() -> None:
    args = parse_args()
    source = pd.read_excel(args.source).reset_index(drop=True)
    source.insert(0, "global_index", source.index)
    source.insert(1, "original_excel_row", source["global_index"] + 2)
    source["PART_ID"] = source["global_index"] // 100 + 1
    source["row_in_part"] = source["global_index"] % 100

    results = read_output_files(args.part_folder, source_rows=len(source))
    best = results.drop_duplicates("global_index", keep="first")

    merged = source.merge(
        best[["global_index", DISTANCE_COL, "_source_output_file"]],
        on="global_index",
        how="left",
    ).rename(columns={"_source_output_file": "source_output_file"})

    parsed = merged[DISTANCE_COL].apply(parse_distance_time)
    merged[KM_COL] = parsed.apply(lambda item: item[0])
    merged[TIME_COL] = parsed.apply(lambda item: item[1])

    text = merged[DISTANCE_COL].astype("string").str.strip()
    missing = merged[DISTANCE_COL].isna() | text.isna() | text.eq("")
    error = text.str.contains(ERROR_RE, na=False)
    failed = missing | error

    merged["crawl_status"] = "ok"
    merged.loc[missing, "crawl_status"] = "missing_distance"
    merged.loc[error, "crawl_status"] = "not_found_or_error"

    remaining = merged.loc[failed].copy()

    args.final.parent.mkdir(parents=True, exist_ok=True)
    args.remaining.parent.mkdir(parents=True, exist_ok=True)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    merged.to_excel(args.final, index=False)
    remaining.to_excel(args.remaining, index=False)

    wb = load_workbook(args.final)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    km_idx = headers.index(KM_COL) + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=km_idx).number_format = "0.00"
    wb.save(args.final)

    summary_lines = [
        f"source_rows={len(source)}",
        f"output_rows_before_dedup={len(results)}",
        f"output_rows_after_dedup={len(best)}",
        f"final_rows={len(merged)}",
        f"ok_rows={len(merged) - len(remaining)}",
        f"remaining_failed_rows={len(remaining)}",
        f"final_file={args.final}",
        f"remaining_file={args.remaining}",
    ]
    args.summary.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    print("\n".join(summary_lines))


if __name__ == "__main__":
    main()
